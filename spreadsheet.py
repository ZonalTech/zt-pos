"""Excel (.xlsx) helpers for bulk import + downloadable templates.

Used by the Products and Purchase-order import flows. Reading also accepts CSV
so a plainly-saved file still works. Pure-Python (openpyxl), so it bundles fine
with PyInstaller.
"""
import csv
import io

from openpyxl import Workbook, load_workbook


def build_template(headers, sample_row=None, sheet_title="Template"):
    """Return .xlsx bytes with a bold header row and one example row."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(headers)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    if sample_row:
        ws.append(sample_row)
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(14, len(h) + 4)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def read_table(file_storage):
    """Parse an uploaded .xlsx or .csv into a list of dicts keyed by header.

    Headers are lower-cased and trimmed. Blank rows are skipped. Returns
    ``(rows, error)`` — error is a user-facing string, or None on success.
    """
    name = (file_storage.filename or "").lower()
    try:
        if name.endswith(".csv"):
            text = file_storage.read().decode("utf-8-sig", errors="replace")
            reader = csv.reader(io.StringIO(text))
            table = [list(r) for r in reader]
        else:
            wb = load_workbook(file_storage, read_only=True, data_only=True)
            ws = wb.active
            table = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    except Exception as e:
        return None, f"Couldn't read the file: {e}"

    if not table:
        return [], None
    headers = [str(h).strip().lower() if h is not None else "" for h in table[0]]
    rows = []
    for raw in table[1:]:
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        row = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            row[h] = raw[i] if i < len(raw) else None
        rows.append(row)
    return rows, None
